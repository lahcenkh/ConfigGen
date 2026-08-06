import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from configgen.core.bulk import read_csv_rows, read_rows, read_xlsx_rows, run_bulk
from configgen.core.schema import Field, Schema
from configgen.hooks import Services

TS = datetime(2026, 1, 15, 14, 30, 0)


def _schema(**overrides) -> Schema:
    defaults = dict(
        name="Widget",
        id="widget",
        template="widget.j2",
        identity_field="name",
        fields=[
            Field(key="name", label="Name", type="string", required=True),
        ],
    )
    defaults.update(overrides)
    return Schema(**defaults)


def _write_template(tmp_path: Path, text: str = "hello {{ name }}") -> Path:
    template_path = tmp_path / "widget.j2"
    template_path.write_text(text, encoding="utf-8")
    return template_path


# -- row readers ---------------------------------------------------------


def test_read_csv_rows(tmp_path: Path):
    csv_path = tmp_path / "rows.csv"
    csv_path.write_text("name,site\nedge-01,nyc\nedge-02,lon\n", encoding="utf-8")
    rows = read_csv_rows(csv_path)
    assert rows == [{"name": "edge-01", "site": "nyc"}, {"name": "edge-02", "site": "lon"}]


def test_read_csv_rows_handles_utf8_bom(tmp_path: Path):
    csv_path = tmp_path / "rows.csv"
    csv_path.write_bytes("name\nedge-01\n".encode("utf-8-sig"))
    rows = read_csv_rows(csv_path)
    assert rows == [{"name": "edge-01"}]
    assert "﻿name" not in rows[0]


def test_read_xlsx_rows(tmp_path: Path):
    xlsx_path = tmp_path / "rows.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["name", "site"])
    sheet.append(["edge-01", "nyc"])
    sheet.append(["edge-02", "lon"])
    workbook.save(xlsx_path)

    rows = read_xlsx_rows(xlsx_path)
    assert rows == [{"name": "edge-01", "site": "nyc"}, {"name": "edge-02", "site": "lon"}]


def test_read_xlsx_rows_skips_trailing_blank_rows(tmp_path: Path):
    xlsx_path = tmp_path / "rows.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["name"])
    sheet.append(["edge-01"])
    sheet.append([None])
    workbook.save(xlsx_path)

    rows = read_xlsx_rows(xlsx_path)
    assert rows == [{"name": "edge-01"}]


def test_read_rows_dispatches_by_extension(tmp_path: Path):
    csv_path = tmp_path / "rows.csv"
    csv_path.write_text("name\nedge-01\n", encoding="utf-8")
    xlsx_path = tmp_path / "rows.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["name"])
    workbook.active.append(["edge-02"])
    workbook.save(xlsx_path)

    assert read_rows(csv_path) == [{"name": "edge-01"}]
    assert read_rows(xlsx_path) == [{"name": "edge-02"}]


# -- run_bulk: basics ---------------------------------------------------------


def test_run_bulk_all_valid_rows_generates_and_manifests(tmp_path: Path):
    _write_template(tmp_path)
    schema = _schema()
    rows = [{"name": "web01"}, {"name": "web02"}]

    result = run_bulk(
        schema,
        rows,
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        username="alice",
        timestamp=TS,
    )

    assert result.valid_count == 2
    assert result.error_count == 0
    assert result.output_dir.name == "batch_20260115143000"
    assert result.manifest_path.is_file()

    txt_files = sorted(result.output_dir.glob("*.txt"))
    assert len(txt_files) == 2
    assert txt_files[0].read_text(encoding="utf-8").endswith("hello web01")

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["batch_id"] == result.batch_id
    assert manifest["valid_count"] == 2
    assert manifest["error_count"] == 0
    assert manifest["source"] == "test.csv"
    assert len(manifest["rows"]) == 2


def test_run_bulk_partial_success_reports_row_errors(tmp_path: Path):
    _write_template(tmp_path)
    schema = _schema()
    rows = [{"name": "web01"}, {"name": ""}, {"name": "web03"}]

    result = run_bulk(
        schema,
        rows,
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        timestamp=TS,
    )

    assert result.valid_count == 2
    assert result.error_count == 1
    assert result.row_errors[0].row_number == 3  # header=1, first data row=2
    assert "name" in result.row_errors[0].errors

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["errors"][0]["row_number"] == 3


def test_run_bulk_zero_valid_rows_still_writes_manifest(tmp_path: Path):
    _write_template(tmp_path)
    schema = _schema()
    rows = [{"name": ""}]

    result = run_bulk(
        schema,
        rows,
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        timestamp=TS,
    )

    assert result.valid_count == 0
    assert result.error_count == 1
    assert result.output_dir.is_dir()
    assert result.manifest_path.is_file()


def test_run_bulk_each_row_gets_its_own_filename(tmp_path: Path):
    _write_template(tmp_path)
    schema = _schema()
    rows = [{"name": "web01"}, {"name": "web02"}]

    result = run_bulk(
        schema,
        rows,
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        timestamp=TS,
    )
    filenames = {doc for row in result.generated for doc in row["documents"].values()}
    assert len(filenames) == 2
    assert any("web01" in f for f in filenames)
    assert any("web02" in f for f in filenames)


# -- run_bulk: hook integration ---------------------------------------


def test_run_bulk_with_hook_success(tmp_path: Path):
    (tmp_path / "widget.j2").write_text("hello {{ cfg_name }}", encoding="utf-8")
    (tmp_path / "hook.py").write_text(
        "def build(values, context, services):\n"
        "    return {'cfg_name': values['name'].upper()}\n",
        encoding="utf-8",
    )
    schema = _schema(hook="hook")

    result = run_bulk(
        schema,
        [{"name": "web01"}],
        templates_dir=tmp_path,
        hooks_dir=tmp_path,
        output_root=tmp_path / "out",
        source="test.csv",
        services=Services(),
        timestamp=TS,
    )
    assert result.valid_count == 1
    text = next(result.output_dir.glob("*.txt")).read_text(encoding="utf-8")
    assert text.endswith("hello WEB01")


def test_run_bulk_hook_error_becomes_row_error_not_a_crash(tmp_path: Path):
    (tmp_path / "widget.j2").write_text("hello {{ cfg_name }}", encoding="utf-8")
    (tmp_path / "hook.py").write_text(
        "from configgen.hooks import HookError\n"
        "\n"
        "def build(values, context, services):\n"
        "    if values['name'] == 'ghost':\n"
        "        raise HookError({'name': 'unknown device'})\n"
        "    return {'cfg_name': values['name']}\n",
        encoding="utf-8",
    )
    schema = _schema(hook="hook")

    result = run_bulk(
        schema,
        [{"name": "web01"}, {"name": "ghost"}],
        templates_dir=tmp_path,
        hooks_dir=tmp_path,
        output_root=tmp_path / "out",
        source="test.csv",
        services=Services(),
        timestamp=TS,
    )
    assert result.valid_count == 1
    assert result.error_count == 1
    assert result.row_errors[0].errors == {"name": "unknown device"}


def test_run_bulk_hook_schema_without_services_raises_valueerror(tmp_path: Path):
    _write_template(tmp_path, "hello {{ cfg_name }}")
    schema = _schema(hook="hook")
    with pytest.raises(ValueError):
        run_bulk(
            schema,
            [{"name": "web01"}],
            templates_dir=tmp_path,
            hooks_dir=None,
            output_root=tmp_path / "out",
            source="test.csv",
            timestamp=TS,
        )


# -- run_bulk: from_db field validation ---------------------------------------


class _FakeDatabase:
    def __init__(self, tables: dict[str, list[str]]):
        self._tables = tables

    def all(self, query_name):
        return self._tables[query_name]


def test_run_bulk_validates_from_db_choice_field(tmp_path: Path):
    _write_template(tmp_path, "region {{ region }}")
    schema = _schema(
        fields=[
            Field(key="region", label="Region", type="choice", from_db={"query": "regions"}),
        ],
        identity_field=None,
    )
    database = _FakeDatabase({"regions": ["us-east", "us-west"]})

    result = run_bulk(
        schema,
        [{"region": "us-east"}, {"region": "mars"}],
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        database=database,
        timestamp=TS,
    )
    assert result.valid_count == 1
    assert result.error_count == 1
    assert "region" in result.row_errors[0].errors


# -- run_bulk: render errors ---------------------------------------------------


def test_run_bulk_render_error_becomes_row_error(tmp_path: Path):
    # Template references a variable no field or hook provides.
    (tmp_path / "widget.j2").write_text("{{ missing_var }}", encoding="utf-8")
    schema = _schema()

    result = run_bulk(
        schema,
        [{"name": "web01"}],
        templates_dir=tmp_path,
        hooks_dir=None,
        output_root=tmp_path / "out",
        source="test.csv",
        timestamp=TS,
    )
    assert result.valid_count == 0
    assert result.error_count == 1
    assert "_render" in result.row_errors[0].errors


# -- run_bulk: logging ---------------------------------------------------------


def test_run_bulk_logs_batch_start_and_summary(tmp_path: Path, caplog):
    _write_template(tmp_path)
    schema = _schema()

    with caplog.at_level("INFO", logger="configgen.core.bulk"):
        result = run_bulk(
            schema,
            [{"name": "web01"}, {"name": ""}],
            templates_dir=tmp_path,
            hooks_dir=None,
            output_root=tmp_path / "out",
            source="rows.csv",
            timestamp=TS,
        )

    messages = "\n".join(caplog.messages)
    assert f"bulk {result.batch_id}: schema=widget rows=2 source=rows.csv" in messages
    assert f"bulk {result.batch_id}: row 3 failed validation" in messages
    assert f"bulk {result.batch_id}: finished — 1 valid, 1 errors" in messages
